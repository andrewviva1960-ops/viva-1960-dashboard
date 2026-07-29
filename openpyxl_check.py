import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
path = r"C:\Users\Andro\Downloads\Financial Model\Viva Financial model 2026 (3)_FIXED.xlsx"

# Try reading with openpyxl directly (formulas, not cached values)
wb = load_workbook(path, data_only=False)
ws = wb["Expenses Raw Data 2026"]

print("=== Expense sheet: first 10 rows, columns E-G (date, Month, Quarters) ===")
for row_idx in range(1, 11):
    e = ws.cell(row=row_idx, column=5).value  # date
    f = ws.cell(row=row_idx, column=6).value  # Month
    g = ws.cell(row=row_idx, column=7).value  # Quarters
    print(f"  Row {row_idx}: date={repr(e)}, Month={repr(f)}, Quarters={repr(g)}")

print("\n=== Expense sheet: check ALL rows for Month values ===")
has_month = 0
for row in ws.iter_rows(min_row=1, max_col=6, values_only=True):
    if row[5] is not None and str(row[5]).strip() != "":
        has_month += 1
        if has_month <= 5:
            print(f"  Found Month={repr(row[5])} in row")
print(f"Total rows with Month value: {has_month}")

# Also check the PNL Dashboard for expense values with data_only=True (cached values)
wb2 = load_workbook(path, data_only=True)
ws2 = wb2["PNL Dashboard "]

print("\n=== PNL Dashboard: Expenses Summary area (rows 32-42, col 1-3) ===")
for row_idx in range(32, 43):
    a = ws2.cell(row=row_idx, column=1).value
    b = ws2.cell(row=row_idx, column=2).value
    c = ws2.cell(row=row_idx, column=3).value
    print(f"  Row {row_idx}: {repr(a)}, {repr(b)}, {repr(c)}")

# Check if monthly expense values exist anywhere in PNL
print("\n=== PNL Dashboard: Scan for any numeric values in columns 2-15 ===")
for row_idx in range(31, 50):
    for col_idx in range(2, 16):
        v = ws2.cell(row=row_idx, column=col_idx).value
        if v is not None and v != "":
            try:
                num = float(v)
                if num == num:  # not NaN
                    print(f"  [{row_idx},{col_idx}] = {num}")
            except:
                pass
